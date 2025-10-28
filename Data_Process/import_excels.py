
# -*- coding: utf-8 -*-
import re
import sys
import argparse
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from filter_db import (
    init_db, insert_test, parse_hms_to_seconds, safe_float
)

# ---------- Helpers to find values on messy forms ----------

def norm(s: Any) -> str:
    """Normalize a cell's text for matching."""
    if s is None:
        return ""
    s = str(s)
    return re.sub(r"\s+", " ", s).strip().lower()

def first_number_in(s: Any) -> Optional[float]:
    """Extract the first float-looking number from a string."""
    if s is None:
        return None
    txt = str(s).replace(",", "")
    m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", txt)
    return float(m.group(0)) if m else None

def find_value(ws, patterns: List[str]) -> Optional[Any]:
    """
    Search whole sheet for a label that matches any regex in patterns.
    If found at (r,c), try to read a numeric/text value from:
      - same row, next columns
      - next row, same column
    Return the raw value (not yet coerced to float); None if not found.
    """
    compiled = [re.compile(pat) for pat in patterns]
    for r in ws.iter_rows(values_only=False):
        for cell in r:
            text = norm(cell.value)
            if any(p.search(text) for p in compiled):
                # 1) same row, to the right
                row = cell.row
                col = cell.column
                for cc in range(col + 1, col + 6):
                    v = ws.cell(row=row, column=cc).value
                    if v not in (None, ""):
                        return v
                # 2) next row, same column
                v = ws.cell(row=row + 1, column=col).value
                if v not in (None, ""):
                    return v
    return None

def find_section_numbers(ws, section_patterns: List[str], lookahead_rows: int = 25) -> List[float]:
    """
    Find a section header (e.g., 'depth measurements') and then collect
    numeric values in the next few rows (first number in each row).
    """
    comp = [re.compile(p) for p in section_patterns]
    for r in ws.iter_rows(values_only=False):
        for cell in r:
            if any(c.search(norm(cell.value)) for c in comp):
                row0 = cell.row + 1
                out: List[float] = []
                for rr in range(row0, min(row0 + lookahead_rows, ws.max_row + 1)):
                    row_vals = [ws.cell(row=rr, column=cc).value for cc in range(1, ws.max_column + 1)]
                    # Take the first numeric on the row
                    num = None
                    for v in row_vals:
                        fv = first_number_in(v)
                        if fv is not None:
                            num = fv
                            break
                    if num is not None:
                        out.append(num)
                    else:
                        # stop when we hit an empty/non-numeric line
                        if out:
                            return out
                return out
    return []

def parse_time_value(v: Any) -> Optional[int]:
    """
    Accepts Excel time cells or strings:
      - '2:29:33', '2.29.33', '01:15'
      - Excel time/datetime cells are converted to string fallback.
    Returns seconds (int) or None.
    """
    if v is None or str(v).strip() == "":
        return None
    # If it's already a string like 2.29.33 or 2:29:33:
    if isinstance(v, str):
        return parse_hms_to_seconds(v)
    # Sometimes Excel stores time as number (fraction of a day)
    try:
        from datetime import time, datetime, timedelta
        if hasattr(v, "hour"):  # time or datetime
            if hasattr(v, "second"):
                return v.hour * 3600 + v.minute * 60 + v.second
        # numeric days -> seconds
        f = float(v)
        if 0 <= f < 2:
            return int(round(f * 24 * 3600))
    except Exception:
        pass
    # fallback: try as string
    return parse_hms_to_seconds(str(v))

# ---------- Field patterns ----------

P = {
    "test_name":        [r"^test\s*name"],
    "date_iso":         [r"^date"],
    "mains_pressure":   [r"mains\s*pressure"],
    "sample_name":      [r"^sample"],
    "cloth_filter":     [r"cloth\s*filter"],
    "pre_notes":        [r"pre-?test\s*notes"],
    "post_notes":       [r"post-?test\s*notes"],

    "dry_solids_g":     [r"dry\s*solids"],
    "water_g":          [r"\bwater\b"],
    "percent_solids":   [r"percent\s*solids|%[\s-]*solids"],

    # Filling
    "fill_pressure_bar":[r"^filling.*pressure|^pressure\s*:\s*$"],
    "fill_time":        [r"^filling.*time|^time\s*:\s*$"],
    "fill_filtrate_g":  [r"^filling.*filtrate"],

    # Cake
    "cake_depth_mm":    [r"^cake.*depth|^depth\s*:\s*$"],
    "cake_area_m2":     [r"surface\s*area"],
    "cake_wet_weight_g":[r"\bweight\b(?!.*moisture)|cake.*weight"],
    "cake_moisture_pct":[r"moisture\s*content"],

    # Pressing
    "press_pressure_bar":[r"^pressing.*pressure"],
    "press_time":        [r"^pressing.*time"],
    "press_filtrate_g":  [r"^pressing.*filtrate"],

    # Air blow
    "air_blow_pressure_bar":[r"air\s*blow.*pressure"],
    "air_blow_time":        [r"air\s*blow.*time"],
    "air_blow_filtrate_g":  [r"air\s*blow.*filtrate"],

    "solids_density_kg_m3":[r"solids\s*density"],
}

SECTION_DEPTHS = [r"depth\s*measurements"]
SECTION_MOIST   = [r"moisture\s*content(?!.*cake)"]

def extract_record(ws) -> Tuple[Dict[str, Any], List[float], List[float]]:
    """Pull a single test record + depth replicates + moisture replicates from a worksheet."""
    rec: Dict[str, Any] = {}

    # Simple fields
    test_name = find_value(ws, P["test_name"])
    date_raw  = find_value(ws, P["date_iso"])
    mains_p   = find_value(ws, P["mains_pressure"])

    sample    = find_value(ws, P["sample_name"])
    cloth     = find_value(ws, P["cloth_filter"])
    pre_n     = find_value(ws, P["pre_notes"])
    post_n    = find_value(ws, P["post_notes"])

    dry_g     = find_value(ws, P["dry_solids_g"])
    water_g   = find_value(ws, P["water_g"])
    pct_sol   = find_value(ws, P["percent_solids"])

    fill_p    = find_value(ws, P["fill_pressure_bar"])
    fill_t    = find_value(ws, P["fill_time"])
    fill_f    = find_value(ws, P["fill_filtrate_g"])

    cake_d    = find_value(ws, P["cake_depth_mm"])
    cake_a    = find_value(ws, P["cake_area_m2"])
    cake_w    = find_value(ws, P["cake_wet_weight_g"])
    cake_m    = find_value(ws, P["cake_moisture_pct"])

    press_p   = find_value(ws, P["press_pressure_bar"])
    press_t   = find_value(ws, P["press_time"])
    press_f   = find_value(ws, P["press_filtrate_g"])

    air_p     = find_value(ws, P["air_blow_pressure_bar"])
    air_t     = find_value(ws, P["air_blow_time"])
    air_f     = find_value(ws, P["air_blow_filtrate_g"])

    rho_s     = find_value(ws, P["solids_density_kg_m3"])

    # Replicate sections
    depths    = find_section_numbers(ws, SECTION_DEPTHS)
    moistures = find_section_numbers(ws, SECTION_MOIST)

    # Coerce & build record matching DB columns
    rec = {
        "test_name": str(test_name) if test_name else None,
        "date_iso":  str(date_raw) if date_raw else None,
        "mains_pressure_bar": safe_float(mains_p),

        "sample_name": str(sample) if sample else None,
        "cloth_filter": str(cloth) if cloth else None,

        "pre_notes":  str(pre_n) if pre_n else "",
        "post_notes": str(post_n) if post_n else "",

        "dry_solids_g": safe_float(dry_g),
        "water_g": safe_float(water_g),
        "percent_solids": safe_float(pct_sol),

        "fill_pressure_bar": safe_float(fill_p),
        "fill_time_s": parse_time_value(fill_t),
        "fill_filtrate_g": safe_float(fill_f),

        "cake_depth_mm": safe_float(cake_d),
        "cake_area_m2": safe_float(cake_a),
        "cake_wet_weight_g": safe_float(cake_w),
        "cake_moisture_pct": safe_float(cake_m),

        "press_pressure_bar": safe_float(press_p),
        "press_time_s": parse_time_value(press_t),
        "press_filtrate_g": safe_float(press_f),

        "air_blow_pressure_bar": safe_float(air_p),
        "air_blow_time_s": parse_time_value(air_t),
        "air_blow_filtrate_g": safe_float(air_f),

        "solids_density_kg_m3": safe_float(rho_s),
    }

    return rec, depths, moistures

# ---------- CLI ----------

def import_folder(folder: Path, db_path: Path, sheet_name: Optional[str] = None) -> None:
    init_db(str(db_path))
    files = sorted([p for p in folder.glob("**/*") if p.suffix.lower() in (".xlsx", ".xlsm")])
    if not files:
        print(f"No Excel files found under: {folder}")
        return

    total = 0
    for f in files:
        try:
            wb = load_workbook(filename=str(f), data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

            rec, depths, moistures = extract_record(ws)

            # basic sanity: require at least a test name or date; skip empty
            if not any([rec.get("test_name"), rec.get("date_iso")]):
                print(f"Skip (no key fields): {f.name}")
                continue

            test_id = insert_test(str(db_path), rec, depths_mm=depths, moistures_pct=moistures)
            print(f"OK: {f.name}  -> test_id={test_id} (depths {len(depths)}, moist {len(moistures)})")
            total += 1
        except Exception as e:
            print(f"ERROR parsing {f.name}: {e}")

    print(f"\nImported {total} file(s) into {db_path}")

def main():
    ap = argparse.ArgumentParser(description="Import press-filter Excel sheets into SQLite.")
    ap.add_argument("folder", help="Folder containing Excel files (.xlsx/.xlsm)")
    ap.add_argument("--db", default="filter_tests.db", help="SQLite DB path (default: filter_tests.db)")
    ap.add_argument("--sheet", default=None, help="Worksheet name to use (default: active sheet)")
    args = ap.parse_args()

    import_folder(Path(args.folder), Path(args.db), sheet_name=args.sheet)

if __name__ == "__main__":
    main()
